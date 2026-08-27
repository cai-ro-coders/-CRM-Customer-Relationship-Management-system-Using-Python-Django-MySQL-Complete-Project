import csv
import io
from datetime import timedelta
from decimal import Decimal

from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone

from crm.decorators import crm_login_required
from crm.models import (Customer, Lead, Deal, Task, Meeting, Invoice, Payment, Ticket,
                        Product, Quote, User, InvoiceItem)


def _range(request):
    days = int(request.GET.get('days', 90))
    end = timezone.now()
    start = end - timedelta(days=days)
    return start, end


def _xl_cell(v):
    """Numbers stay numbers in Excel; everything else is text."""
    if isinstance(v, (int, float, Decimal)):
        return v
    s = str(v).replace(',', '').replace('$', '').strip()
    try:
        return Decimal(s)
    except Exception:
        return str(v)


def _export_response(request, header, rows, name):
    from django.utils.html import escape
    fmt = request.GET.get('format', 'csv')
    title = name.replace('_', ' ').title()
    if fmt == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]
        ws.append([str(h) for h in header])
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='198754')
        for row in rows:
            ws.append([_xl_cell(v) for v in row])
        for i, h in enumerate(header, 1):
            width = max([len(str(h))] + [len(str(r[i - 1])) for r in rows]) + 2 if rows else len(str(h)) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(width, 50)
        ws.freeze_panes = 'A2'
        buf = io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{name}.xlsx"'
    elif fmt == 'pdf':
        # print-friendly inline page; browser prints / saves as PDF
        html = ['<html><head><meta charset="utf-8">',
                f'<title>{title}</title>',
                '<style>',
                'body{font-family:Arial,Helvetica,sans-serif;margin:24px;color:#111}',
                'h2{margin:0 0 4px} .meta{color:#555;font-size:12px;margin-bottom:16px}',
                'table{border-collapse:collapse;width:100%;font-size:12px}',
                'th,td{border:1px solid #999;padding:6px 8px;text-align:left}',
                'th{background:#f0f0f0} tr:nth-child(even) td{background:#fafafa}',
                '@page{size:A4 landscape;margin:14mm}',
                '@media print{.noprint{display:none}}',
                '</style></head><body>',
                f'<h2>{title}</h2>',
                f'<div class="meta">{timezone.now().strftime("%d %b %Y %H:%M")} · '
                f'{len(rows)} records</div>',
                '<button class="noprint" onclick="window.print()" '
                'style="padding:8px 16px;margin-bottom:12px;cursor:pointer">Print / Save as PDF</button>',
                '<table><tr>' + ''.join(f'<th>{escape(h)}</th>' for h in header) + '</tr>']
        for row in rows:
            html.append('<tr>' + ''.join(f'<td>{escape(str(v))}</td>' for v in row) + '</tr>')
        html.append('</table>')
        html.append('<script>window.addEventListener("load",function(){setTimeout(function(){window.print()},300)});</script>')
        html.append('</body></html>')
        resp = HttpResponse(''.join(html), content_type='text/html; charset=utf-8')
        resp['Content-Disposition'] = f'inline; filename="{name}.pdf"'
    else:
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(header)
        w.writerows(rows)
        resp = HttpResponse(buf.getvalue(), content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{name}.{fmt}"'
    return resp


@crm_login_required
def reports_home(request):
    today = timezone.now().date()
    stats = {
        'customers': Customer.objects.filter(is_deleted=False).count(),
        'leads': Lead.objects.count(),
        'open_deals': Deal.objects.filter(status='Open').count(),
        'won_value': Deal.objects.filter(status='Won').aggregate(s=Sum('expected_value'))['s'] or 0,
        'invoices': Invoice.objects.count(),
        'payments': Payment.objects.aggregate(s=Sum('amount'))['s'] or 0,
        'open_tickets': Ticket.objects.exclude(status='Closed').count(),
        'tasks_pending': Task.objects.exclude(status__in=['Completed', 'Cancelled']).count(),
    }
    return render(request, 'reports/home.html', {'stats': stats})


@crm_login_required
def sales_report(request):
    start, end = _range(request)
    qs = Invoice.objects.select_related('customer').filter(invoice_date__gte=start.date())
    header = ['Invoice #', 'Date', 'Customer', 'Total', 'Paid', 'Balance', 'Status']
    rows = [[i.invoice_number, i.invoice_date.strftime('%Y-%m-%d'), i.customer.company_name,
             str(i.total), str(i.paid_amount), str(i.balance), i.status] for i in qs]
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'sales_report')
    return render(request, 'reports/table.html', {
        'title': 'Sales Report',
        'header': header,
        'rows': rows,
        'total': sum(float(r[3]) for r in rows),
        'days': request.GET.get('days', 90),
    })


@crm_login_required
def customer_report(request):
    qs = Customer.objects.filter(is_deleted=False).select_related('assigned_to')
    header = ['Code', 'Company', 'Contact', 'Email', 'City', 'Country', 'Deals',
              'Invoiced', 'Status']
    rows = []
    for c in qs:
        invoiced = Invoice.objects.filter(customer=c).aggregate(s=Sum('total'))['s'] or 0
        deals_n = Deal.objects.filter(customer=c).count()
        rows.append([c.customer_code, c.company_name, c.full_name, c.email, c.city, c.country,
                     deals_n, str(invoiced), c.status])
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'customer_report')
    return render(request, 'reports/table.html', {
        'title': 'Customer Report', 'header': header, 'rows': rows, 'total': len(rows),
        'days': None})


@crm_login_required
def lead_report(request):
    qs = Lead.objects.select_related('source', 'status', 'assigned_to')
    header = ['Lead', 'Company', 'Source', 'Status', 'Priority', 'Expected Revenue',
              'Assigned To', 'Created']
    rows = [[l.full_name, l.company, l.source.name if l.source else '-',
             l.status.name if l.status else '-', l.priority, str(l.expected_revenue),
             l.assigned_to.full_name if l.assigned_to else '-',
             l.created_at.strftime('%Y-%m-%d')] for l in qs]
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'lead_report')
    by_source = list(Lead.objects.values('source__name').annotate(c=Count('id')))
    return render(request, 'reports/lead.html', {
        'title': 'Lead Report', 'header': header, 'rows': rows, 'by_source': by_source,
        'total_revenue': sum(float(r[5]) for r in rows)})


@crm_login_required
def revenue_report(request):
    start, end = _range(request)
    payments = (Payment.objects.select_related('invoice', 'customer')
                .filter(payment_date__gte=start.date()))
    header = ['Date', 'Invoice #', 'Customer', 'Method', 'Reference', 'Amount']
    rows = [[p.payment_date.strftime('%Y-%m-%d'), p.invoice.invoice_number if p.invoice else '-',
             p.customer.company_name, p.payment_method, p.transaction_reference or '-',
             str(p.amount)] for p in payments]
    total = sum(p.amount for p in payments)
    monthly = list(Payment.objects.extra(select={'m': "DATE_FORMAT(payment_date, '%%Y-%%m')"})
                   .values('m').annotate(total=Sum('amount')).order_by('m'))
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'revenue_report')
    return render(request, 'reports/revenue.html', {
        'title': 'Revenue Report', 'header': header, 'rows': rows, 'total': total,
        'monthly': monthly, 'days': str(request.GET.get('days', '90'))})


@crm_login_required
def invoice_report(request):
    header = ['Invoice #', 'Date', 'Due Date', 'Customer', 'Total', 'Paid', 'Balance', 'Status']
    rows = [[i.invoice_number, i.invoice_date.strftime('%Y-%m-%d'),
             i.due_date.strftime('%Y-%m-%d') if i.due_date else '-',
             i.customer.company_name, str(i.total), str(i.paid_amount), str(i.balance), i.status]
            for i in Invoice.objects.select_related('customer')]
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'invoice_report')
    overdue_total = sum(i.balance for i in Invoice.objects.filter(status='Overdue'))
    return render(request, 'reports/table.html', {
        'title': 'Invoice Report', 'header': header, 'rows': rows,
        'extra_stat': f'Outstanding (Overdue): {overdue_total:,.2f}', 'days': None})


@crm_login_required
def payment_report(request):
    header = ['Date', 'Invoice', 'Customer', 'Method', 'Amount']
    rows = [[p.payment_date.strftime('%Y-%m-%d'), p.invoice.invoice_number if p.invoice else '-',
             p.customer.company_name, p.payment_method, str(p.amount)]
            for p in Payment.objects.select_related('invoice', 'customer')]
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'payment_report')
    return render(request, 'reports/table.html', {
        'title': 'Payment Report', 'header': header, 'rows': rows,
        'total': sum(float(r[4]) for r in rows), 'days': None})


@crm_login_required
def product_report(request):
    header = ['SKU', 'Product', 'Category', 'Price', 'Tax %', 'Stock', 'Units Sold', 'Revenue']
    from django.db.models import F
    sold = (InvoiceItem.objects.values('product_id', 'product__product_name', 'product__sku',
                                       'product__category__category_name', 'product__unit_price',
                                       'product__tax_rate', 'product__stock_quantity')
            .annotate(units=Sum('quantity'), revenue=Sum(F('quantity') * F('price'))))
    rows = [[s['product__sku'], s['product__product_name'],
             s['product__category__category_name'] or '-',
             str(s['product__unit_price']), str(s['product__tax_rate']),
             s['product__stock_quantity'], int(s['units'] or 0), str(s['revenue'] or 0)]
            for s in sold]
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'product_report')
    return render(request, 'reports/table.html', {
        'title': 'Product Report', 'header': header, 'rows': rows, 'days': None})


@crm_login_required
def ticket_report(request):
    header = ['Ticket #', 'Subject', 'Customer', 'Category', 'Priority', 'Status', 'Replies',
              'Created']
    rows = [[t.ticket_number, t.subject, t.customer.company_name,
             t.category.name if t.category else '-', t.priority, t.status,
             t.replies.count(), t.created_at.strftime('%Y-%m-%d')]
            for t in Ticket.objects.select_related('customer', 'category')]
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'ticket_report')
    by_status = list(Ticket.objects.values('status').annotate(c=Count('id')))
    return render(request, 'reports/ticket.html', {
        'title': 'Support Report', 'header': header, 'rows': rows, 'by_status': by_status})


@crm_login_required
def employee_report(request):
    header = ['User', 'Role', 'Customers', 'Leads', 'Open Deals', 'Won Deals',
              'Won Value', 'Tasks Pending']
    rows = []
    for u in User.objects.select_related('role'):
        rows.append([u.full_name or u.username, u.role.name if u.role else '-',
                     Customer.objects.filter(assigned_to=u).count(),
                     Lead.objects.filter(assigned_to=u).count(),
                     Deal.objects.filter(assigned_to=u, status='Open').count(),
                     Deal.objects.filter(assigned_to=u, status='Won').count(),
                     str(Deal.objects.filter(assigned_to=u, status='Won').aggregate(
                         s=Sum('expected_value'))['s'] or 0),
                     Task.objects.filter(assigned_to=u).exclude(
                         status__in=['Completed', 'Cancelled']).count()])
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'employee_report')
    return render(request, 'reports/table.html', {
        'title': 'Employee Performance Report', 'header': header, 'rows': rows, 'days': None})


@crm_login_required
def task_report(request):
    header = ['Title', 'Assigned To', 'Related Customer', 'Priority', 'Status', 'Due Date']
    rows = [[t.title, t.assigned_to.full_name if t.assigned_to else '-',
             t.customer.company_name if t.customer else '-', t.priority, t.status,
             t.due_date.strftime('%Y-%m-%d') if t.due_date else '-']
            for t in Task.objects.select_related('assigned_to', 'customer')]
    if request.GET.get('format'):
        return _export_response(request, header, rows, 'task_report')
    by_status = list(Task.objects.values('status').annotate(c=Count('id')))
    return render(request, 'reports/task.html', {
        'title': 'Task Report', 'header': header, 'rows': rows, 'by_status': by_status})
